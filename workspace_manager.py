"""
Workspace management for SF Wizard
Handles Excel file organization and workspace operations
"""

import os
import shutil
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter
from excel_generator import generate_where_in


class WorkspaceManager:
    """Manages workspaces and Excel files"""
    
    def __init__(self, excel_files_dir: str = "excel-files"):
        self.excel_files_dir = Path(excel_files_dir)
        self.excel_files_dir.mkdir(exist_ok=True)
    
    def list_workspaces(self) -> list:
        """List all existing workspaces"""
        workspaces = []
        
        if not self.excel_files_dir.exists():
            return workspaces
        
        for workspace_dir in self.excel_files_dir.iterdir():
            if workspace_dir.is_dir():
                current_files = list((workspace_dir / "current").glob("*.xlsx"))
                if current_files:
                    workspaces.append({
                        "name": workspace_dir.name,
                        "file": current_files[0].name
                    })
        
        return sorted(workspaces, key=lambda x: x["name"])
    
    def get_workspace_info(self, workspace: str) -> dict:
        """Get workspace information"""
        workspace_dir = self.excel_files_dir / workspace
        
        if not workspace_dir.exists():
            raise ValueError(f"Workspace '{workspace}' not found")
        
        current_files = list((workspace_dir / "current").glob("*.xlsx"))
        versions = list((workspace_dir / "versions").glob("*_v*.xlsx"))

        latest_version_file = None
        if versions:
            # pick the version with highest vN in filename if possible, otherwise by mtime
            def extract_v(n):
                m = re.search(r'_v(\d+)(?:_|\.|$)', n.name)
                return int(m.group(1)) if m else -1

            versions_sorted = sorted(versions, key=lambda f: extract_v(f) if extract_v(f) >= 0 else f.stat().st_mtime, reverse=True)
            latest_version_file = versions_sorted[0].name

        return {
            "name": workspace,
            "current_file": current_files[0].name if current_files else None,
            "version_count": len(versions),
            "latest_version_file": latest_version_file,
        }
    
    def create_or_update_workspace(self, workspace_name: str, file_content: bytes) -> str:
        """
        Create a new workspace or update existing one
        Returns the workspace name
        """
        workspace_dir = self.excel_files_dir / workspace_name
        
        # Create directory structure if new workspace
        is_new = not workspace_dir.exists()
        if is_new:
            workspace_dir.mkdir(parents=True, exist_ok=True)
            (workspace_dir / "original").mkdir(exist_ok=True)
            (workspace_dir / "current").mkdir(exist_ok=True)
            (workspace_dir / "versions").mkdir(exist_ok=True)
        
        # Save original file (only on first upload)
        if is_new:
            original_file = workspace_dir / "original" / f"{workspace_name}.xlsx"
            with open(original_file, "wb") as f:
                f.write(file_content)

        # Save as current file
        current_dir = workspace_dir / "current"

        # If a current file exists, move it to versions (version it)
        existing_current = list(current_dir.glob("*.xlsx"))
        if existing_current:
            # compute next version and move existing current to versions
            ver = self.get_next_version_number(workspace_name)
            src = existing_current[0]
            dst = workspace_dir / "versions" / f"{workspace_name}_v{ver}.xlsx"
            src.rename(dst)

        # Write new current
        current_file = current_dir / f"{workspace_name}.xlsx"
        with open(current_file, "wb") as f:
            f.write(file_content)
        
        return workspace_name
    
    def get_excel_columns(self, workspace: str) -> list:
        """Get available columns from the current Excel file"""
        workspace_dir = self.excel_files_dir / workspace
        current_files = list((workspace_dir / "current").glob("*.xlsx"))
        
        if not current_files:
            raise ValueError(f"No Excel file in workspace '{workspace}'")
        
        excel_file = current_files[0]
        wb = load_workbook(excel_file)

        columns = []
        seen = set()

        # We'll track sheet columns that are covered by named tables to avoid duplicates
        exempt_sheet_cols = set()

        # Process named tables first
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # openpyxl stores tables in ws.tables (dict)
            try:
                tables = list(ws.tables.values())
            except Exception:
                tables = []

            for table in tables:
                # table.name, table.tableColumns (list of TableColumn)
                table_name = getattr(table, 'name', None) or getattr(table, 'displayName', None)
                cols = []
                for tc in getattr(table, 'tableColumns', []) or []:
                    # TableColumn has .name
                    colname = getattr(tc, 'name', None)
                    if colname:
                        entry = f"{table_name}[{colname}]"
                        if entry not in seen:
                            columns.append(entry)
                            seen.add(entry)
                # Mark the sheet columns covered by this table as exempt
                ref = getattr(table, 'ref', None)
                if ref:
                    try:
                        min_col, min_row, max_col, max_row = range_boundaries(ref)
                        for c in range(min_col, max_col + 1):
                            exempt_sheet_cols.add((sheet_name, get_column_letter(c)))
                    except Exception:
                        pass

        # Now add sheet-wide columns (Sheet!A:A), skipping exempted ones
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            max_col = ws.max_column or 0
            for c in range(1, max_col + 1):
                col_letter = get_column_letter(c)
                if (sheet_name, col_letter) in exempt_sheet_cols:
                    continue
                entry = f"{sheet_name}!{col_letter}:{col_letter}"
                if entry not in seen:
                    columns.append(entry)
                    seen.add(entry)

        return columns
    
    def get_next_version_number(self, workspace: str) -> int:
        """Get the next version number for a workspace"""
        versions_dir = self.excel_files_dir / workspace / "versions"
        
        if not versions_dir.exists():
            return 1
        
        max_version = 0
        for file in versions_dir.glob("*_v*.xlsx"):
            # Extract version number from filename
            match = re.search(r'_v(\d+)(?:_|\.|$)', file.name)
            if match:
                version = int(match.group(1))
                max_version = max(max_version, version)
        
        return max_version + 1
    
    def generate_where_in(
        self,
        workspace: str,
        generator_name: str,
        source: str,
        where_col: str,
        has_header: bool = False,
        soql_base: str = ""
    ) -> str:
        """
        Generate a WHERE IN query in the Excel file
        Returns the version filename
        """
        workspace_dir = self.excel_files_dir / workspace
        current_files = list((workspace_dir / "current").glob("*.xlsx"))
        
        if not current_files:
            raise ValueError(f"No Excel file in workspace '{workspace}'")
        
        input_file = current_files[0]
        version_num = self.get_next_version_number(workspace)
        output_filename = f"{workspace}_v{version_num}_{generator_name}.xlsx"
        output_file = workspace_dir / "versions" / output_filename
        
        # Generate the WHERE IN query
        generate_where_in(
            input_path=str(input_file),
            output_path=str(output_file),
            gen_name=generator_name,
            source_ref=source,
            has_header=has_header,
            where_col=where_col,
            soql_base=soql_base
        )
        
        return output_filename
    
    @staticmethod
    def _get_column_letter(col_num: int) -> str:
        """Convert column number to letter (1 -> A, 27 -> AA)"""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(65 + col_num % 26) + result
            col_num //= 26
        return result
